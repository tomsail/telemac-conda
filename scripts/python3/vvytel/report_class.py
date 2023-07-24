r"""@author TELEMAC-MASCARET Consortium

    @note ... this work is based on a collaborative effort between
                 TELEMAC-MASCARET consortium members
    @brief Report class for validation
"""
# _____          ___________________________________________________
# ____/ Imports /__________________________________________________/
#
# ~~> dependencies towards standard python
import time as ttime
from os import path, sep
from collections import OrderedDict
from utils.exceptions import TelemacException
from config import CFGS


def insert_key(odict, key, after=None):
    """
    Insert a key in an OrderedDict either after a specific key or it assumes
    the keys are sorted and places it were it belongs

    @param odict (OrderedDict) Dict in which to insert key
    @param key (str) Key to insert
    @param after (str) If not None the key is added after 'after' if after is
    '' it is added in first position
    """

    odict.update({key:OrderedDict()})

    if after is not None:

        move = after == ''
        for okey in list(odict.keys())[:-1]:
            if move:
                odict.move_to_end(okey)
            else:
                if okey == after:
                    move = True
    else:
        for okey in list(odict.keys())[:-1]:
            if okey > key:
                odict.move_to_end(okey)

    return odict

def get_value(sheet, cell):
    """
    Get value of a cell that is with a merged cell

    @param sheet (openpyxl.worksheet) Sheet to look in
    @param cell (openpyxl.worksheet.cell) cell to look for
    """
    def within_range(bounds, cell):
        column_start, row_start, column_end, row_end = bounds
        row = cell.row
        if row >= row_start and row <= row_end:
            column = cell.column
            if column >= column_start and column <= column_end:
                return True
        return False

    for merged in sheet.merged_cells:
        if within_range(merged.bounds, cell):
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value

def plot_bar_report_time(data, title, fig_name=''):
    """
    Plot of stat from report

    @param data (dict) Dictionary from compute stat
    @param title (str) Title of the figure
    @param fig_name (str) If given save figure instead of displaying it
    """
    import matplotlib.pyplot as plt
    import numpy as np

    pre_times = [item['pre'] for item in data.values()]
    run_times = [item['run'] for item in data.values()]
    vnv_times = [item['vnv'] for item in data.values()]
    post_times = [item['post'] for item in data.values()]

    ind = np.arange(len(data))
    width = 0.35

    p_pre = plt.bar(ind, pre_times, width)
    p_run = plt.bar(ind, run_times, width, bottom=pre_times)
    p_vnv = plt.bar(ind, vnv_times, width, bottom=np.add(pre_times, run_times))
    bottom = np.add(np.add(pre_times, run_times), vnv_times)
    p_post = plt.bar(ind, post_times, width, bottom=bottom)

    plt.ylabel('times (s)')
    plt.title(title)
    plt.xticks(ind, data.keys(), rotation=90, ha="center")
    plt.legend((p_pre[0], p_run[0], p_vnv[0], p_post[0]),
               ('pre', 'run', 'vnv', 'post'))

    if fig_name != '':
        plt.savefig(fig_name)
    else:
        plt.show()

    plt.close('all')

def get_report_path(report_name, type_valid):
    """
    Build report full path
    $HOMETEL/[report_name]_[config]_[version]_[type_valid]_[date].csv
    where:
    - report_name is the one given as argument
    - config is the name of the configuration for which the validation is run
    - version is the version of the code
    - type_valid is the one given as argument
    - date is the data at which the report is written

    @param report_name (str) Name given to the report
    @param type_valid (str) Type of validation (notebook, examples...)


    @returns (str) The path
    """

    full_report_name = "{}_{}_{}_{}_{}.csv".format(\
                            report_name,
                            CFGS.cfgname,
                            CFGS.configs[CFGS.cfgname].get('version', 'main'),
                            type_valid,
                            ttime.strftime("%Y-%m-%d-%Hh%Mmin%Ss",
                                           ttime.localtime(ttime.time())))

    report_path = path.join(CFGS.get_root(), full_report_name)

    return report_path

class Report():
    """
    Reader/writer for validation report
    """

    def __init__(self, name, type_valid, full_name=False):
        """

        @param name (str) Name of the report (full_path will be build from
        there)
        @param type_valid (str) Type of validation (examples, notebooks..)
        name
        """

        if type_valid not in ['examples', 'notebooks']:
            raise TelemacException(\
                    "Unknown validation type: {}".format(type_valid))

        self.type_valid = type_valid
        self.values = OrderedDict()
        if name != '':
            self.file_name = get_report_path(name, type_valid)
        else:
            self.file_name = ''

    def add_notebook(self, nb_file, time, passed):
        """
        Adding a notebook to the report

        @param nb_file (str) Name of the notebook
        @param time (float) execution time
        @param passed (bool) If the notebook worked
        """
        if self.type_valid != 'notebooks':
            raise TelemacException(\
                    'add_action is only possible for a notebooks report')

        self.values[nb_file] = {\
                'time':time,
                'passed':passed}

    def add_action(self, file_name, rank, action, time, passed):
        """
        Add a new action to the report

        @param file_name (str) Name of the file for which the action
        @param rank (int) Rank of the action
        @param action (str) Name of the action
        @param time (float) Time to run the action
        @param passed (bool) If the action worked
        """
        if self.type_valid != 'examples':
            raise TelemacException(\
                    'add_action is only possible for an example report')

        if file_name not in self.values:
            self.values[file_name] = OrderedDict()

        self.values[file_name][action] = {\
                'rank':rank,
                'time':time,
                'passed':passed}

    def read(self, file_name=None):
        """
        Read data from existing file

        @param file_name (str) Name of the file to read from (by default will
        use the name from class)
        """
        if file_name is None:
            file_name = self.file_name

        with open(file_name, 'r') as f:
            header = f.readline().split(';')
            if header[0] == 'Notebook file':
                self.type_valid = 'notebooks'
                for line in f:
                    nb_file, time, passed = line.split(';')
                    self.add_notebook(nb_file, float(time), 'True' in passed)
            else:
                self.type_valid = 'examples'
                for line in f:
                    py_file, rank, action, time, passed = line.split(';')
                    self.add_action(py_file, int(rank), action, float(time),
                                    'True' in passed)


    def write(self, file_name=None):
        """
        Write content of class into a file

        @param file_name (str) Name of the output file (by default will take
        name used with Report was initialised)
        """

        if file_name is None:
            file_name = self.file_name

        if self.type_valid == 'examples':
            header = "Python file;rank;action_name;duration;passed\n"
            with open(file_name, 'w') as f:
                f.write(header)
                for ffile_name, actions in self.values.items():
                    for action_name, action_info in actions.items():
                        llist = [ffile_name,
                                 str(action_info['rank']),
                                 action_name,
                                 str(action_info['time']),
                                 str(action_info['passed'])]
                        f.write(';'.join(llist)+'\n')
        elif self.type_valid == 'notebooks':
            header = "Notebook file;duration;passed\n"
            with open(file_name, 'w') as f:
                f.write(header)
                for ffile_name, actions in self.values.items():
                    llist = [ffile_name,
                             str(actions['time']),
                             str(actions['passed'])]
                    f.write(';'.join(llist)+'\n')

    def write2xls(self, xls_file, mode, title, job_id, date, verbose=False, max_report=0):
        """
        Transfer data from report_file into an xls file

        @param xls_file (str) Path to the xls file
        @param mode (str) mode to write append, create, insert
        @param title (str) Title of the worsheet
        @param job_id (str) Id for the data column
        @param date (str) Date for the data column
        @param verbose (str) If true more verbose
        @param max_report (str) Max of report within the file (delete older report)
        """
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils.cell import get_column_letter
        from openpyxl.styles import PatternFill, Alignment, Font

        if self.type_valid != "examples":
            raise TelemacException("write2xls only avaialable for examples type")

        if mode != 'create':
            wb = load_workbook(filename=xls_file)
        else:
            wb = Workbook()

        ws = wb.worksheets[0]

        if mode == 'create':
            if verbose:
                print("Creating header for fixed columns")
            ws.title = title

            ws['A3'] = "Module"
            ws['A3'].font = Font(bold=True)
            ws['B3'] = "Case"
            ws['B3'].font = Font(bold=True)
            ws['C3'] = "Python File"
            ws['C3'].font = Font(bold=True)
            ws['D3'] = "Rank"
            ws['D3'].font = Font(bold=True)
            ws['E3'] = "action_name"
            ws['E3'].font = Font(bold=True)
            # Freezing column and row of headers
            ws.freeze_panes = ws['F4']


        green_color = "0000FF00"
        red_color = "00FF0000"


        # Get first and last column
        # Index of first report column
        first_column_index = 6
        # index of first empty report column
        last_column_index = first_column_index

        if verbose:
            print("Identifying last column position")
        cell = ws['{}1'.format(get_column_letter(last_column_index))]
        while cell.value is not None:
            last_column_index += 2
            col_let = get_column_letter(last_column_index)
            cell = ws['{}1'.format(col_let)]


        # define were the new columns will be added
        if mode == 'insert':
            column_index = first_column_index
            if verbose:
                print("Adding new columns")
            ws.insert_cols(column_index)
            ws.insert_cols(column_index)
        elif mode == 'apppend':
            column_index = last_column_index
        else:
            column_index = first_column_index

        last_column_index += 2

        if max_report > 0:
            if (last_column_index - first_column_index)//2 > max_report:
                if mode == 'append':
                    # delete first two columns
                    ws.delete_cols(first_column_index)
                    ws.delete_cols(first_column_index)
                    last_column_index -= 2
                if mode == 'insert':
                    # delete last two columns
                    ws.delete_cols(last_column_index-2)
                    ws.delete_cols(last_column_index-2)
                    last_column_index -= 2


        start_line = 4

        if verbose:
            print("Adding new columns header")
        # Writting job number
        column_letter1 = get_column_letter((column_index))
        ws['{}3'.format(column_letter1)] = "time (s)"
        ws['{}3'.format(column_letter1)].font = Font(bold=True)

        column_letter2 = get_column_letter((column_index+1))
        ws['{}3'.format(column_letter2)] = "passed"
        ws['{}3'.format(column_letter2)].font = Font(bold=True)

        # Merge job number cells
        ws.merge_cells('{}1:{}1'.format(column_letter1, column_letter2))
        ws['{}1'.format(column_letter1)] = "{} ({})".format(job_id, date)
        ws['{}1'.format(column_letter1)].alignment = \
                Alignment(horizontal="center", vertical="center")
        ws['{}1'.format(column_letter1)].font = Font(bold=True)

        # Number of passed and failed
        ws['{}2'.format(column_letter1)].fill = \
                    PatternFill(start_color=green_color,
                                end_color=green_color,
                                fill_type="solid")

        ws['{}2'.format(column_letter2)].fill = \
                    PatternFill(start_color=red_color,
                                end_color=red_color,
                                fill_type="solid")

        # update of all sums (Did not manage to create relative sums the insert
        # mode break them)
        for column in range(first_column_index, last_column_index, 2):

            column_letter_1 = get_column_letter(column)
            column_letter_2 = get_column_letter(column+1)
            ws['{}2'.format(column_letter_1)] = \
                    '=COUNTIF({col}{start}:{col}{end},"passed")'\
                    .format(col=column_letter_2, start=start_line, end=3000)
            ws['{}2'.format(column_letter_1)].font = Font(bold=True)
            ws['{}2'.format(column_letter_2)] = \
                    '=COUNTIF({col}{start}:{col}{end},"failed")'\
                    .format(col=column_letter_2, start=start_line, end=3000)
            ws['{}2'.format(column_letter_2)].font = Font(bold=True)

            if mode == 'insert':
                ws.merge_cells('{}1:{}1'.format(column_letter_1, column_letter_2))

        # Unmerging A through E column (issues when append)
        if mode != 'create':
            if verbose:
                print("Unmerging cells")
            # Merging cell with same name for
            # Module, case, py_file, rank (same merge as py_file)
            for column in ['A', 'B', 'C', 'D']:
                first_line = start_line
                iline = start_line + 1
                for iline in range(start_line+1, ws.max_row+1):
                    val = ws['{}{}'.format(column, iline)].value
                    if val is not None or iline == ws.max_row:
                        if iline == ws.max_row:
                            b_range = first_line
                            end_line = ws.max_row
                        else:
                            b_range = first_line+1
                            end_line = iline-1
                        ws.unmerge_cells('{0}{1}:{0}{2}'\
                            .format(column, first_line, end_line))
                        # Copy back what was in merged cell into other cells
                        for mg_line in range(b_range, end_line+1):
                            ws['{}{}'.format(column, mg_line)].value = \
                                    ws['{}{}'.format(column, first_line)].value
                        first_line = iline

        # building dict strucre of module, cases, py_file, action
        # existing combinaison to detect new inputs (new module, case,
        # py_file, action)
        used = OrderedDict()
        if verbose:
            print("Building dict structure")
        if mode != 'create':
            for line in range(start_line, ws.max_row+1):
                mod = ws['A{}'.format(line)].value
                if mod not in used:
                    used[mod] = OrderedDict()
                case = ws['B{}'.format(line)].value
                if case not in used[mod]:
                    used[mod][case] = OrderedDict()
                py_file = ws['C{}'.format(line)].value
                if py_file not in used[mod][case]:
                    used[mod][case][py_file] = OrderedDict()
                action = ws['E{}'.format(line)].value
                used[mod][case][py_file][action] = line
            last_line = line

            # check for difference between input file and excel structure
            # And add new entries in dict
            stuff_added = False
            for file_name, actions in self.values.items():
                head, tail = path.split(file_name)
                py_file = tail
                head, tail = path.split(head)
                case = tail
                head, tail = path.split(head)
                mod = tail
                # Adding new dict if not already in
                if mod not in used:
                    insert_key(used, mod)
                    stuff_added = True
                if case not in used[mod]:
                    insert_key(used[mod], case)
                    stuff_added = True
                if py_file not in used[mod][case]:
                    insert_key(used[mod][case], py_file)
                    stuff_added = True
                prev_action = ''
                # Actions are not sort by name so adding them after previous one
                for action in actions:
                    if action not in used[mod][case][py_file]:
                        insert_key(used[mod][case][py_file], action,
                                   after=prev_action)
                        used[mod][case][py_file][action] = 0
                        stuff_added = True
                    prev_action = action

            # Update line info and create new lines for the new entries
            if stuff_added:
                line = start_line
                new_lines = 0
                for mod in used:
                    for case in used[mod]:
                        for py_file in used[mod][case]:
                            for action in used[mod][case][py_file]:
                                res = used[mod][case][py_file][action]
                                # New line
                                if res == 0:
                                    ws.insert_rows(line)
                                    used[mod][case][py_file][action] = line
                                    new_lines += 1
                                    line += 1
                                else:
                                    # Updating old line value
                                    used[mod][case][py_file][action] += \
                                            new_lines
                                    line = used[mod][case][py_file][action] + 1
                last_line = line

        else:
            # Building line number for each report line
            line = start_line
            for file_name, actions in self.values.items():
                head, tail = path.split(file_name)
                py_file = tail
                head, tail = path.split(head)
                case = tail
                head, tail = path.split(head)
                mod = tail
                if mod not in used:
                    used[mod] = OrderedDict()
                if case not in used[mod]:
                    used[mod][case] = OrderedDict()
                if py_file not in used[mod][case]:
                    used[mod][case][py_file] = OrderedDict()
                for action in actions:
                    used[mod][case][py_file][action] = line
                    line += 1
            last_line = line

        for file_name, actions in self.values.items():
            if verbose:
                print("adding element for: ", file_name)
            head, tail = path.split(file_name)
            py_file = tail
            head, tail = path.split(head)
            case = tail
            head, tail = path.split(head)
            module = tail
            for action_name, action_info in actions.items():

                line = used[module][case][py_file][action_name]
                if verbose:
                    print("  adding action for: ", action_name)

                # Module
                ws.cell(row=line, column=1).value = module
                # Case
                ws.cell(row=line, column=2).value = case
                # python file
                ws.cell(row=line, column=3).value = py_file
                # rank
                ws.cell(row=line, column=4).value = action_info['rank']
                # action_name
                ws.cell(row=line, column=5).value = action_name

                # time
                ws.cell(row=line, column=column_index).value = action_info['time']
                # passed
                cell = ws.cell(row=line, column=column_index+1)
                cell.value = "passed" if action_info['passed'] else "failed"
                if action_info['passed']:
                    color = "0000FF00"
                else:
                    color = "00FF0000"
                cell.fill = PatternFill(start_color=color,
                                        end_color=color,
                                        fill_type="solid")

        # Merging cell with same name for
        # Module, case, py_file, rank (same merge as py_file)
        if verbose:
            print("mergin cells")
        for column in ['A', 'B', 'C']:
            first_line = start_line
            merge_val = ws['{}{}'.format(column, first_line)].value
            for iline in range(start_line+1, last_line+2):
                val = ws['{}{}'.format(column, iline)].value
                if val != merge_val:
                    if first_line < iline:
                        ws.merge_cells('{column}{first_line}:{column}{iline}'\
                                       .format(column=column,
                                               first_line=first_line,
                                               iline=iline-1))
                        ws['{}{}'.format(column, first_line)].alignment = \
                                Alignment(horizontal="center", vertical="center")
                        # Merge rank column the same way as py_file column
                        if column == 'C':
                            ws.merge_cells('D{first_line}:D{iline}'\
                                           .format(column=column,
                                                   first_line=first_line,
                                                   iline=iline-1))
                            ws['D{}'.format(first_line)].alignment = \
                                    Alignment(horizontal="center", vertical="center")
                    first_line = iline
                    merge_val = val


        # Resizing columns
        if verbose:
            print("Resizing columns")
        factor = 1.2
        for idx, col in enumerate(ws.columns, 1):
            vals = (len(u"{0}".format(c.value)) for c in col)
            max_width = max(vals) * factor
            ws.column_dimensions[get_column_letter(idx)].width = max_width

        # Setting filters
        if verbose:
            print("Adding filters")
        col_let = get_column_letter(ws.max_column-1)
        ws.auto_filter.ref = "A3:{}{}".format(col_let, ws.max_row)
        #ws.auto_filter.add_filter_column(0, list(used.keys()))

        if verbose:
            print("Saving file")
        wb.save(filename=xls_file)

    def compute_stats(self):
        """
        Will compute stats from the report info
        """
        if self.type_valid != "examples":
            raise TelemacException(
               "Compute_stats only available for examples type report")

        module_time = {}
        rank_time = {}
        per_module_time = {}

        for py_file, actions in self.values.items():
            module = get_module_from_path(py_file)
            if module not in module_time:
                module_time[module] = \
                        {'pre':0.0, 'run':0.0, 'vnv':0.0, 'post':0.0}
                per_module_time[module] = {}
            short_name = path.basename(py_file)
            per_module_time[module][short_name] = \
                        {'pre':0.0, 'run':0.0, 'vnv':0.0, 'post':0.0}
            for action, data in actions.items():
                rank = data['rank']
                time = data['time']
                if rank not in rank_time:
                    rank_time[rank] = \
                            {'pre':0.0, 'run':0.0, 'vnv':0.0, 'post':0.0}
                if action not in ['pre', 'vnv', 'post']:
                    per_module_time[module][short_name]['run'] += time
                    module_time[module]['run'] += time
                    rank_time[rank]['run'] += time
                else:
                    per_module_time[module][short_name][action] += time
                    module_time[module][action] += time
                    rank_time[rank][action] += time

        return per_module_time, module_time, rank_time

    def plot_stats(self):
        """
        Plot stat from
        """
        per_module_time, module_time, rank_time = self.compute_stats()

        plot_bar_report_time(\
                 module_time,
                 'Execution time per module',
                 fig_name='module_time.png')
        plot_bar_report_time(\
                rank_time,
                'Execution time per rank',
                fig_name='rank_time.png')

        for module, time in per_module_time.items():
            plot_bar_report_time(\
                    time,
                    'Exuction time per vnv_*.py for '+module,
                    fig_name=module+'_time.png')


def get_module_from_path(file_name):
    """
    extract module from path
    """
    names = file_name.split(sep)
    i = names.index('examples')
    return names[i+1]
